#!/usr/bin/env python3
"""
ManicTime Data Import Script

Imports ManicTime Excel data into SQLite database.
Supports multiple Excel files with different sheet structures.
"""

import os
import sys
import sqlite3
from datetime import datetime
from pathlib import Path

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent.parent))

try:
    import pandas as pd
    import openpyxl
except ImportError:
    print("Error: Required packages not installed")
    print("Please install: pip install pandas openpyxl")
    sys.exit(1)


class ManicTimeImporter:
    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = None
        self.cursor = None

    def connect(self):
        """Connect to SQLite database"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()

        # Enable WAL mode
        self.cursor.execute('PRAGMA journal_mode=WAL')

        print(f"Connected to database: {self.db_path}")

    def create_tables(self):
        """Create tables if they don't exist"""
        migration_file = Path(__file__).parent / 'migrations' / '003_create_manictime_tables.sql'

        if migration_file.exists():
            with open(migration_file, 'r', encoding='utf-8') as f:
                sql = f.read()
                self.cursor.executescript(sql)
                self.conn.commit()
                print("Tables created successfully")
        else:
            print(f"Warning: Migration file not found: {migration_file}")

    def parse_excel_file(self, excel_path):
        """Parse ManicTime Excel/CSV file with automatic format detection"""
        print(f"\n=== Processing: {excel_path} ===")

        # Try to detect file type
        is_csv = excel_path.lower().endswith('.csv')

        try:
            if is_csv:
                # Try reading as CSV first
                try:
                    print("Attempting to read as CSV...")
                    df = pd.read_csv(excel_path, encoding='utf-8')
                    print(f"CSV read successful!")
                    print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")
                    print(f"  Columns: {list(df.columns)[:5]}")

                    # Process the CSV data
                    activities = self.parse_sheet(df, "Sheet1")
                    return activities

                except UnicodeDecodeError:
                    # Try different encodings
                    for encoding in ['gbk', 'gb2312', 'latin1', 'cp1252']:
                        try:
                            print(f"Trying encoding: {encoding}")
                            df = pd.read_csv(excel_path, encoding=encoding)
                            print(f"Success with {encoding}!")
                            print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")
                            activities = self.parse_sheet(df, "Sheet1")
                            return activities
                        except:
                            continue

                    # If all CSV attempts fail, try as Excel
                    print("CSV read failed, trying as Excel...")
                    is_csv = False

            if not is_csv:
                # Try reading as Excel with openpyxl
                try:
                    print("Attempting to read as Excel with openpyxl...")
                    from openpyxl import load_workbook
                    wb = load_workbook(excel_path, read_only=True, data_only=True)

                    activities = []
                    for sheet_name in wb.sheetnames:
                        print(f"\nProcessing sheet: {sheet_name}")
                        ws = wb[sheet_name]

                        # Convert to DataFrame manually
                        data = []
                        headers = None
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i == 0:
                                headers = row
                            else:
                                data.append(row)

                        if headers and data:
                            df = pd.DataFrame(data, columns=headers)
                            print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")
                            print(f"  Columns: {list(df.columns)[:5]}")
                            activities_from_sheet = self.parse_sheet(df, sheet_name)
                            activities.extend(activities_from_sheet)

                    wb.close()
                    return activities

                except Exception as e:
                    print(f"openpyxl read failed: {e}")

                    # Try with pandas read_excel
                    try:
                        print("Trying pandas read_excel...")
                        xl = pd.ExcelFile(excel_path)
                        activities = []

                        for sheet_name in xl.sheet_names:
                            print(f"\nProcessing sheet: {sheet_name}")
                            df = pd.read_excel(excel_path, sheet_name=sheet_name)
                            print(f"  Rows: {len(df)}, Columns: {len(df.columns)}")
                            activities_from_sheet = self.parse_sheet(df, sheet_name)
                            activities.extend(activities_from_sheet)

                        return activities

                    except Exception as e2:
                        print(f"pandas read_excel also failed: {e2}")

            return []

        except Exception as e:
            print(f"Error reading file: {e}")
            return []

    def parse_sheet(self, df, sheet_name):
        """Parse a single sheet and extract activities"""
        activities = []

        # Common ManicTime column patterns
        time_cols = ['Start time', 'End time', 'Start', 'End', 'From', 'To']
        app_cols = ['Application', 'App', 'Program', 'Process']
        duration_cols = ['Duration', 'Time', 'Length']
        title_cols = ['Title', 'Window', 'Document', 'Window title']

        # Find matching columns
        start_col = None
        end_col = None
        app_col = None
        duration_col = None
        title_col = None

        for col in df.columns:
            col_lower = str(col).lower()

            if any(t.lower() in col_lower for t in time_cols) and 'start' in col_lower:
                start_col = col
            elif any(t.lower() in col_lower for t in time_cols) and 'end' in col_lower:
                end_col = col
            elif any(a.lower() in col_lower for a in app_cols):
                app_col = col
            elif any(d.lower() in col_lower for d in duration_cols):
                duration_col = col
            elif any(t.lower() in col_lower for t in title_cols):
                title_col = col

        print(f"  Detected columns: start={start_col}, end={end_col}, app={app_col}, duration={duration_col}")

        if not app_col:
            print(f"  Skipping sheet {sheet_name}: No application column found")
            return activities

        # Process rows
        for idx, row in df.iterrows():
            try:
                app = str(row[app_col]) if pd.notna(row[app_col]) else None

                if not app or app == 'nan':
                    continue

                # Get start and end times
                start_time = None
                end_time = None
                duration_seconds = 0

                if start_col and pd.notna(row[start_col]):
                    # Handle Unix timestamps (integers) or datetime strings
                    start_val = row[start_col]
                    if isinstance(start_val, (int, float)):
                        # Unix timestamp
                        start_time = pd.to_datetime(start_val, unit='s')
                    else:
                        start_time = pd.to_datetime(start_val)

                if end_col and pd.notna(row[end_col]):
                    # Handle Unix timestamps (integers) or datetime strings
                    end_val = row[end_col]
                    if isinstance(end_val, (int, float)):
                        # Unix timestamp
                        end_time = pd.to_datetime(end_val, unit='s')
                    else:
                        end_time = pd.to_datetime(end_val)

                # Calculate duration
                if start_time and end_time:
                    duration_seconds = int((end_time - start_time).total_seconds())
                elif duration_col and pd.notna(row[duration_col]):
                    # Try to parse duration
                    duration_val = row[duration_col]
                    if isinstance(duration_val, (int, float)):
                        duration_seconds = int(duration_val)
                    elif isinstance(duration_val, pd.Timedelta):
                        duration_seconds = int(duration_val.total_seconds())

                if duration_seconds <= 0:
                    continue

                # Get window title
                title = None
                if title_col and pd.notna(row[title_col]):
                    title = str(row[title_col])[:500]  # Limit length

                # Get date
                date = None
                if start_time:
                    date = start_time.strftime('%Y%m%d')

                activity = {
                    'date': date,
                    'start_time': start_time.isoformat() if start_time else None,
                    'end_time': end_time.isoformat() if end_time else None,
                    'duration_seconds': duration_seconds,
                    'application': app[:200],  # Limit length
                    'window_title': title,
                    'category': self.categorize_application(app)
                }

                activities.append(activity)

            except Exception as e:
                print(f"  Error processing row {idx}: {e}")
                continue

        print(f"  Extracted {len(activities)} activities from sheet {sheet_name}")
        return activities

    def categorize_application(self, app_name):
        """Automatically categorize application"""
        app_lower = app_name.lower()

        # Development
        if any(x in app_lower for x in ['vscode', 'visual studio', 'pycharm', 'intellij', 'eclipse', 'sublime', 'atom', 'vim', 'emacs', 'code']):
            return 'Development'

        # Browsers
        if any(x in app_lower for x in ['chrome', 'firefox', 'edge', 'safari', 'brave', 'opera']):
            return 'Browser'

        # Communication
        if any(x in app_lower for x in ['slack', 'teams', 'zoom', 'skype', 'discord', 'telegram', 'wechat', 'qq']):
            return 'Communication'

        # Office
        if any(x in app_lower for x in ['word', 'excel', 'powerpoint', 'outlook', 'onenote', 'office']):
            return 'Office'

        # Entertainment
        if any(x in app_lower for x in ['spotify', 'music', 'video', 'youtube', 'netflix', 'game']):
            return 'Entertainment'

        # System
        if any(x in app_lower for x in ['explorer', 'finder', 'terminal', 'cmd', 'powershell', 'bash']):
            return 'System'

        return 'Other'

    def import_activities(self, activities):
        """Import activities into database"""
        if not activities:
            print("No activities to import")
            return

        print(f"\n=== Importing {len(activities)} activities ===")

        # Insert activities
        inserted = 0
        skipped = 0

        for activity in activities:
            try:
                self.cursor.execute('''
                    INSERT INTO manictime_activities
                    (date, start_time, end_time, duration_seconds, application, window_title, category)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    activity['date'],
                    activity['start_time'],
                    activity['end_time'],
                    activity['duration_seconds'],
                    activity['application'],
                    activity['window_title'],
                    activity['category']
                ))
                inserted += 1

                if inserted % 1000 == 0:
                    self.conn.commit()
                    print(f"  Imported {inserted} activities...")

            except sqlite3.IntegrityError:
                skipped += 1
            except Exception as e:
                print(f"  Error inserting activity: {e}")
                skipped += 1

        self.conn.commit()
        print(f"\nImport complete:")
        print(f"  Inserted: {inserted}")
        print(f"  Skipped: {skipped}")

    def update_app_metadata(self):
        """Update application metadata table"""
        print("\n=== Updating application metadata ===")

        self.cursor.execute('''
            INSERT OR REPLACE INTO manictime_apps
            (application, category, first_seen, last_seen, total_duration_seconds, total_sessions)
            SELECT
                application,
                category,
                MIN(date) as first_seen,
                MAX(date) as last_seen,
                SUM(duration_seconds) as total_duration_seconds,
                COUNT(*) as total_sessions
            FROM manictime_activities
            GROUP BY application
        ''')

        self.conn.commit()

        # Get count
        self.cursor.execute('SELECT COUNT(*) FROM manictime_apps')
        count = self.cursor.fetchone()[0]
        print(f"Updated metadata for {count} applications")

    def update_daily_stats(self):
        """Update daily statistics"""
        print("\n=== Updating daily statistics ===")

        self.cursor.execute('''
            INSERT OR REPLACE INTO manictime_daily
            (date, application, total_duration_seconds, session_count)
            SELECT
                date,
                application,
                SUM(duration_seconds) as total_duration_seconds,
                COUNT(*) as session_count
            FROM manictime_activities
            GROUP BY date, application
        ''')

        self.conn.commit()

        # Get count
        self.cursor.execute('SELECT COUNT(*) FROM manictime_daily')
        count = self.cursor.fetchone()[0]
        print(f"Updated {count} daily statistics")

    def print_summary(self):
        """Print import summary"""
        print("\n=== Import Summary ===")

        # Total activities
        self.cursor.execute('SELECT COUNT(*) FROM manictime_activities')
        total_activities = self.cursor.fetchone()[0]
        print(f"Total activities: {total_activities:,}")

        # Total applications
        self.cursor.execute('SELECT COUNT(*) FROM manictime_apps')
        total_apps = self.cursor.fetchone()[0]
        print(f"Total applications: {total_apps}")

        # Date range
        self.cursor.execute('SELECT MIN(date), MAX(date) FROM manictime_activities')
        min_date, max_date = self.cursor.fetchone()
        print(f"Date range: {min_date} to {max_date}")

        # Total duration
        self.cursor.execute('SELECT SUM(duration_seconds) FROM manictime_activities')
        total_seconds = self.cursor.fetchone()[0] or 0
        total_hours = total_seconds / 3600
        print(f"Total duration: {total_hours:.1f} hours")

        # Top 5 apps
        print("\nTop 5 applications:")
        self.cursor.execute('''
            SELECT application, total_duration_seconds, category
            FROM manictime_apps
            ORDER BY total_duration_seconds DESC
            LIMIT 5
        ''')
        for i, (app, duration, category) in enumerate(self.cursor.fetchall(), 1):
            hours = duration / 3600
            print(f"  {i}. {app} ({category}): {hours:.1f}h")

    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
            print("\nDatabase connection closed")


def main():
    # Paths
    script_dir = Path(__file__).parent
    data_dir = script_dir.parent.parent / 'data' / 'screentime'
    manictime_dir = data_dir / 'ManicTime'
    db_path = data_dir / 'manictime_computer.db'

    print("=== ManicTime Data Import ===")
    print(f"Data directory: {manictime_dir}")
    print(f"Database: {db_path}")

    # Check if data directory exists
    if not manictime_dir.exists():
        print(f"Error: Data directory not found: {manictime_dir}")
        sys.exit(1)

    # Find Excel and CSV files
    excel_files = list(manictime_dir.glob('*.xlsx'))
    csv_files = list(manictime_dir.glob('*.csv'))
    all_files = excel_files + csv_files

    if not all_files:
        print(f"Error: No Excel or CSV files found in {manictime_dir}")
        sys.exit(1)

    print(f"\nFound {len(all_files)} data files:")
    for f in all_files:
        size_mb = f.stat().st_size / (1024 * 1024)
        file_type = "CSV" if f.suffix.lower() == '.csv' else "Excel"
        print(f"  - {f.name} ({size_mb:.1f} MB) [{file_type}]")

    # Create importer
    importer = ManicTimeImporter(str(db_path))

    try:
        # Connect and create tables
        importer.connect()
        importer.create_tables()

        # Process each file
        all_activities = []
        for data_file in all_files:
            activities = importer.parse_excel_file(str(data_file))
            all_activities.extend(activities)
            print(f"  Total activities so far: {len(all_activities)}")

        # Import activities
        importer.import_activities(all_activities)

        # Update metadata and statistics
        importer.update_app_metadata()
        importer.update_daily_stats()

        # Print summary
        importer.print_summary()

    except Exception as e:
        print(f"\nError during import: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    finally:
        importer.close()

    print("\n=== Import Complete ===")


if __name__ == '__main__':
    main()
